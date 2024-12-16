import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color

def create_workbook():
    wb = Workbook()
    index_sheet = wb.active
    index_sheet.title = "Index"
    return wb, index_sheet

def process_resource(resource):
    details = resource['Details']
    df_details = pd.json_normalize(details)

    # Process the 'type' field
    resource_type_df = df_details['type'].values[0] if 'type' in df_details.columns else ''
    safe_resource_type_df = resource_type_df.replace('Microsoft.', '').replace('/', '_').lower()

    # Fall back to using ResourceType if type field is empty
    if not safe_resource_type_df:
        resource_type = resource['ResourceType']
        safe_resource_type = resource_type.replace('Microsoft.', '').replace('/', '_').lower()
    else:
        safe_resource_type = safe_resource_type_df

    # Convert the resource type name to a human-friendly format
    parts = safe_resource_type.split('_')
    if len(parts) > 1:
        service_name = parts[0].capitalize()
        resource_type_name = " ".join([part.capitalize() for part in parts[1:]])
        safe_resource_type = f"{service_name} - {resource_type_name}"

    fixed_columns_order = ['type', 'id', 'name']
    for column in fixed_columns_order:
        if column not in df_details.columns:
            df_details[column] = ''

    # Extract subscription id from 'id' column and add it as a new column
    if 'id' in df_details.columns:
        df_details['subscription_id'] = df_details['id'].apply(lambda x: x.split('/')[2] if isinstance(x, str) and len(x.split('/')) > 2 else '')

    # Make 'subscription_id' the first column
    cols = list(df_details.columns)
    cols.insert(0, cols.pop(cols.index('subscription_id')))
    df_details = df_details.loc[:, cols]

    return safe_resource_type, df_details

def update_resource_types(safe_resource_type, df_details, resource_types):
    if len(safe_resource_type) > 31:
        safe_resource_type = safe_resource_type[:31]
    if safe_resource_type not in resource_types:
        other_columns = [col for col in df_details.columns.tolist() if col not in ['type', 'id', 'name', 'subscription_id'] and not col.startswith('tag.')]
        tag_columns = [col for col in df_details.columns.tolist() if col.startswith('tag.')]
        sorted_columns = ['subscription_id', 'type', 'id', 'name'] + sorted(other_columns) + sorted(tag_columns)
        resource_types[safe_resource_type] = sorted_columns
    return safe_resource_type, resource_types

def write_to_workbook(wb: Workbook, safe_resource_type: str, df_details, resource_types):
    if safe_resource_type not in wb.sheetnames:
        wb.create_sheet(title=safe_resource_type)
        wb[safe_resource_type].insert_rows(0)
        # Add a link to the index sheet at the top of the current sheet
        cell = wb[safe_resource_type].cell(row=1, column=1)
        cell.value = "Index"
        cell.hyperlink = "#Index!A1"
        cell.font = Font(color=Color('0563C1'), underline='single')
        for row in dataframe_to_rows(df_details[resource_types[safe_resource_type]], index=False, header=True):
            wb[safe_resource_type].append(row)
            break
    df_details = df_details.reindex(columns=resource_types[safe_resource_type])
    for row in dataframe_to_rows(df_details, index=False, header=False):
        row = [str(cell) if isinstance(cell, list) else cell for cell in row]
        wb[safe_resource_type].append(row)

def output_to_excel(data, output_folder):
    wb, index_sheet = create_workbook()
    resource_types = {}
    recommendations_list = []  # List to collect all recommendations

    for subscription, resource_groups in data['Objects'].items():
        for rg_name, resources in resource_groups.items():
            for resource in resources:
                safe_resource_type, df_details = process_resource(resource)
                safe_resource_type, resource_types = update_resource_types(safe_resource_type, df_details, resource_types)
                write_to_workbook(wb, safe_resource_type, df_details, resource_types)

                # Collect recommendations
                if 'Recommendations' in resource and resource['Recommendations']:
                    for recommendation in resource['Recommendations']:
                        # Add resource context to the recommendation
                        rec = recommendation.copy()
                        rec['SubscriptionId'] = subscription
                        rec['ResourceGroupName'] = rg_name
                        details = resource['Details']

                        # Handle case where 'Details' can be a dict or a list
                        if isinstance(details, dict):
                            rec['ResourceId'] = details.get('id', '')
                            rec['ResourceName'] = details.get('name', '')
                        elif isinstance(details, list):
                            # Use the first item if it's a dict
                            if len(details) > 0 and isinstance(details[0], dict):
                                rec['ResourceId'] = details[0].get('id', '')
                                rec['ResourceName'] = details[0].get('name', '')
                            else:
                                rec['ResourceId'] = ''
                                rec['ResourceName'] = ''
                        else:
                            rec['ResourceId'] = ''
                            rec['ResourceName'] = ''

                        rec['ResourceType'] = resource['ResourceType']
                        recommendations_list.append(rec)

    # Add resource types to the Index sheet with hyperlinks
    for resource_type in resource_types:
        index_sheet.append([resource_type])
        last_row = index_sheet.max_row
        cell = index_sheet.cell(row=last_row, column=1)
        cell.hyperlink = f'#{resource_type}!A1'
        cell.font = Font(color=Color('0563C1'), underline='single')

    # Create the Recommendations sheet if there are any recommendations
    if recommendations_list:
        # Create a DataFrame from the recommendations list
        recommendations_df = pd.json_normalize(recommendations_list)

        # Rearrange columns to bring important ones to the front
        cols_to_front = ['SubscriptionId', 'ResourceGroupName', 'ResourceId', 'ResourceName', 'ResourceType']
        other_columns = [col for col in recommendations_df.columns if col not in cols_to_front]
        recommendations_df = recommendations_df[cols_to_front + other_columns]

        # Create the Recommendations sheet
        wb.create_sheet(title='Recommendations')
        ws = wb['Recommendations']
        # Add a link to the index sheet at the top of the current sheet
        ws.insert_rows(0)
        cell = ws.cell(row=1, column=1)
        cell.value = "Index"
        cell.hyperlink = "#Index!A1"
        cell.font = Font(color=Color('0563C1'), underline='single')

        # Write the DataFrame to the sheet
        for r in dataframe_to_rows(recommendations_df, index=False, header=True):
            ws.append(r)

        # Adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Add the Recommendations sheet to the index
        index_sheet.append(['Recommendations'])
        last_row = index_sheet.max_row
        cell = index_sheet.cell(row=last_row, column=1)
        cell.hyperlink = f'#Recommendations!A1'
        cell.font = Font(color=Color('0563C1'), underline='single')

    # Save the workbook
    wb.save(f'{output_folder}/output.xlsx')
